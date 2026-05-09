# ================= IMPORTS =================
import csv, json, time, ipaddress
import pandas as pd
from pathlib import Path
from threading import Lock
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from nccc_dashboard.config import progress
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from reportlab.platypus import Image, Spacer
import matplotlib.pyplot as plt
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

lock = Lock()

# ================= CONFIG =================
REQUIRED_COLS = [
    "Time","Source IP","Source Country","Source Port",
    "Dstn IP","Unit","Dstn Country","Dstn Port",
    "Protocol","TCP Flag","Pkt Count","Pkt Size"
]

# ================= HELPERS =================
def normalize_ip(ip):
    try:
        if pd.isna(ip):
            return None
        ip = str(ip).strip().split(":")[0]
        return str(ipaddress.ip_address(ip))
    except:
        return None


def detect_branch(p):
    p = str(p).lower()

    if "red" in p:
        return "red"
    elif "white" in p:
        return "white"
    elif "internet" in p or "ids" in p:
        return "ids"
    elif "nkn" in p:
        return "nkn"
    elif "web" in p:
        return "web"
    elif "navy" in p:
        return "navy"
    elif "airforce" in p or "air force" in p:
        return "airforce"

    return "unknown"


def log(msg):
    print(msg, flush=True)
    progress.setdefault("logs", []).append(msg)
    progress["logs"] = progress["logs"][-200:]


def is_valid_file(f):
    name = str(f).lower()
    return not ("application" in name or "malware" in name)


# ================= MYSQL =================
def load_reference_data():
    try:
        import mysql.connector

        conn = mysql.connector.connect(
            host="127.0.0.1",
            user="root",
            password="Root@123",
            database="data1_test",
            use_pure=True
        )

        cur = conn.cursor()
        cur.execute("SELECT ip FROM master_ioc_ip")

        ioc_set = {ip for ip in (normalize_ip(x[0]) for x in cur.fetchall()) if ip}

        unit_map = {}
        for table in ["units","units2","unit3"]:
            try:
                cur.execute(f"SELECT unitIp, unitname FROM {table}")
                for ip,name in cur.fetchall():
                    ip = normalize_ip(ip)
                    if ip:
                        unit_map[ip] = name
            except:
                pass

        conn.close()

        log(f"IOC loaded: {len(ioc_set)}")

        return ioc_set, unit_map

    except Exception as e:
        log(f"MySQL failed: {e}")
        return set(), {}


# ================= CLEAN =================
def clean_csv(inp, out):
    out.parent.mkdir(parents=True, exist_ok=True)

    with open(inp, "r", errors="ignore") as f, open(out, "w", newline="") as w:
        r = csv.reader(f)
        wr = csv.writer(w)

        for row in r:
            if row:
                wr.writerow(row[:16])


# ================= PROCESS FILE =================
def process_file(f, ioc_set, unit_map, out_dir):

    df = pd.read_csv(f, dtype=str, on_bad_lines="skip", low_memory=False)
    df.columns = [c.strip().lower() for c in df.columns]

    col_map = {
        "time":"Time","source ip":"Source IP","destination ip":"Dstn IP",
        "source port":"Source Port","destination port":"Dstn Port",
        "protocol":"Protocol","tcp flag":"TCP Flag",
        "pkt count":"Pkt Count","pkt size":"Pkt Size",
        "source country":"Source Country","destination country":"Dstn Country"
    }

    df = df.rename(columns=col_map)

    for c in REQUIRED_COLS:
        if c not in df.columns:
            df[c] = ""

    df["Source IP"] = df["Source IP"].apply(normalize_ip)
    df["Dstn IP"] = df["Dstn IP"].apply(normalize_ip)

    log(f"{f.name} before filter: {len(df)}")

    if ioc_set:
        df = df[
            df["Source IP"].isin(ioc_set) |
            df["Dstn IP"].isin(ioc_set)
        ]

    log(f"{f.name} after filter: {len(df)}")

    if df.empty:
        return 0

    df["Unit"] = df["Dstn IP"].map(unit_map).fillna("unknown")

    df["Time"] = pd.to_datetime(df["Time"], errors="coerce", dayfirst=True)
    df = df.dropna(subset=["Time"])
    df = df.sort_values("Time")

    df = df[REQUIRED_COLS]

    rel = f.relative_to(out_dir.parent / "csv_cleaned")
    branch = detect_branch(str(f))

    filename = f"{branch}_{rel.parent.name}_{f.stem}_ioc_hits_mail.csv"
    out_file = out_dir / rel.parent / filename
    out_file.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_file, index=False)

    return len(df)


# ================= SUMMARY =================
def build_summary(ioc_dir, json_dir):

    summary = {}

    for f in ioc_dir.rglob("*_ioc_hits_mail.csv"):

        branch = detect_branch(f.name)

        if branch == "unknown":
            continue

        try:
            df = pd.read_csv(f, dtype=str)
        except:
            continue

        if df.empty:
            continue

        df["Time"] = pd.to_datetime(df["Time"], errors="coerce")
        df = df.dropna(subset=["Time"])

        summary[branch] = {
            "total_attacks": len(df),
            "top_ips": df["Source IP"].value_counts().head(10).to_dict(),
            "top_countries": df["Source Country"].value_counts().head(10).to_dict(),
            "protocol_dist": df["Protocol"].value_counts().to_dict(),
            "top_unit": df["Unit"].value_counts().to_dict(),
            "top_dst_ips": df["Dstn IP"].value_counts().head(10).to_dict(),
            "top_src_ports": df["Source Port"].value_counts().head(10).to_dict(),
            "top_dst_ports": df["Dstn Port"].value_counts().head(10).to_dict(),
            "pkt_size_dist": df["Pkt Size"].value_counts().head(10).to_dict(),
            "top_dst_countries": df["Dstn Country"].value_counts().head(10).to_dict(),
            "attacks_by_hour": {
                str(int(k)): int(v)
                for k, v in df["Time"].dt.hour.value_counts().items()
            },
            "timeline": {
                str(k): int(v)
                for k, v in df["Time"].dt.date.value_counts().items()
            }
        }

    # ================= MERGE =================
    merged = {
        "total_attacks": 0,
        "top_ips": {},
        "top_countries": {},
        "protocol_dist": {},
        "top_unit": {},
        "top_dst_ips": {},
        "top_src_ports": {},
        "top_dst_ports": {},
        "pkt_size_dist": {},
        "top_dst_countries": {},
        "attacks_by_hour": {},
        "timeline": {}
    }

    for data in summary.values():

        if not isinstance(data, dict):
            continue

        merged["total_attacks"] += data.get("total_attacks", 0)

        for key in merged.keys():

            if key == "total_attacks":
                continue

            sub = data.get(key, {})

            if isinstance(sub, dict):
                for k, v in sub.items():
                    merged[key][k] = merged[key].get(k, 0) + v

    # ================= SAVE =================
    json_dir.mkdir(parents=True, exist_ok=True)

    # 🔥 IMPORTANT: SAVE ONLY MERGED DATA HERE
    with open(json_dir / "summary_all.json", "w") as f:
        json.dump(merged, f, indent=2)

    # 🔥 SAVE FULL SUMMARY (OPTIONAL)
    with open(json_dir / "summary_full.json", "w") as f:
        json.dump(summary, f, indent=2)

    # 🔥 SAVE BRANCH FILES
    for b in summary:
        with open(json_dir / f"summary_{b}.json", "w") as f:
            json.dump(summary[b], f, indent=2)

    return summary


# ================= DCYA REPORT =================
def build_dcya_report(ioc_dir, dcya_dir):

    dcya_dir.mkdir(parents=True, exist_ok=True)

    TARGET_BRANCHES = ["red", "white", "ids", "nkn", "web"]

    MAX_ROWS = 1_000_000  # safe limit

    # ================= LOAD ALL DATA FIRST =================
    branch_data_map = {}

    for branch in TARGET_BRANCHES:

        log(f"Processing branch: {branch}")

        branch_data = []

        for f in ioc_dir.rglob("*_ioc_hits_mail.csv"):

            fname = f.name.lower()

            if branch == "ids":
                match = ("ids" in fname or "internet" in fname)
            else:
                match = (branch in fname)

            if not match:
                continue

            log(f"Matched file: {f.name}")

            try:
                df = pd.read_csv(f, dtype=str)
            except:
                continue

            if df.empty:
                continue

            df["Time"] = pd.to_datetime(df["Time"], errors="coerce")
            df = df.dropna(subset=["Time"])

            branch_data.append(df)

        if not branch_data:
            log(f"No data for branch: {branch}")
            branch_data_map[branch] = pd.DataFrame()
            continue

        df = pd.concat(branch_data, ignore_index=True)
        df = df.sort_values("Time")

        df["Time"] = df["Time"].dt.strftime("%Y-%m-%d %H:%M:%S")

        branch_data_map[branch] = df

    # ================= SPLIT LOGIC =================
    file_index = 1
    current_rows = 0

    wb = Workbook()
    ws = wb.active
    ws.title = "DCYA"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center")

    columns = REQUIRED_COLS
    row = 1

    def save_file(index):
        out_file = dcya_dir / f"DCYA_REPORT_{index}.xlsx"
        wb.save(out_file)
        log(f"DCYA saved: {out_file}")

    # ================= HEADER =================
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
    ws.cell(row=row, column=1, value="DCYA").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).alignment = center
    row += 2

    # ================= WRITE BRANCHES =================
    for branch in TARGET_BRANCHES:

        df = branch_data_map.get(branch)

        if df is None or df.empty:
            continue

        # 🔥 CHECK IF NEW FILE NEEDED
        if current_rows + len(df) > MAX_ROWS:

            save_file(file_index)

            file_index += 1

            # RESET
            wb = Workbook()
            ws = wb.active
            ws.title = "DCYA"
            row = 1
            current_rows = 0

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
            ws.cell(row=row, column=1, value="DCYA").font = Font(bold=True, size=14)
            ws.cell(row=row, column=1).alignment = center
            row += 2

        # ================= BRANCH TITLE =================
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(columns))
        ws.cell(row=row, column=1, value=branch.upper()).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = center
        row += 1

        # ================= HEADER =================
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = center

        row += 1

        # ================= DATA =================
        for _, r in df.iterrows():

            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row, column=col_idx, value=r.get(col_name, ""))

            row += 1
            current_rows += 1

        row += 2

    # ================= SAVE FINAL =================
    save_file(file_index)


# ================= PDF (EXECUTIVE) =================
# Replaces the old build_pdf_report — drop-in compatible, same signature.
def build_pdf_report(ioc_dir, pdf_dir, json_dir):
    """
    Calls the executive-grade PDF builder from pdf_report.py.
    Falls back to the original basic report if that module is unavailable.
    """
    try:
        from nccc_dashboard.pdf_report import build_pdf_report_v2
        summary_file = json_dir / "summary_all.json"
        result = build_pdf_report_v2(str(summary_file), str(pdf_dir))
        log(f"Executive PDF saved: {result}")
        log("PDF REPORT GENERATED (FINAL)")
    except Exception as e:
        log(f"Executive PDF failed ({e}), falling back to basic report")
        _build_pdf_report_basic(ioc_dir, pdf_dir, json_dir)


def _build_pdf_report_basic(ioc_dir, pdf_dir, json_dir):
    """Original plain PDF — kept as fallback, logic unchanged from original."""

    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors

    pdf_dir.mkdir(parents=True, exist_ok=True)

    styles = getSampleStyleSheet()
    pdf_path = pdf_dir / "SOC_REPORT.pdf"
    doc = SimpleDocTemplate(str(pdf_path))

    content = []

    # ================= HEADER =================
    content.append(Paragraph("NCCC CYBER THREAT REPORT", styles["Title"]))
    content.append(Paragraph(time.strftime("%Y-%m-%d %H:%M:%S"), styles["Normal"]))
    content.append(Spacer(1, 12))

    # ================= LOAD SUMMARY =================
    summary_file = json_dir / "summary_all.json"

    if not summary_file.exists():
        content.append(Paragraph("No summary available", styles["Normal"]))
        doc.build(content)
        log("PDF generated (no summary)")
        return

    with open(summary_file) as f:
        summary = json.load(f)

    # ================= HELPERS =================
    def get_top_attacker_country(d):
        if not isinstance(d, dict) or not d:
            return "-"
        filtered = {k: v for k, v in d.items() if k.upper() != "IN"}
        if not filtered:
            return "IN (Internal)"
        return max(filtered, key=filtered.get)

    def get_top(d):
        return max(d, key=d.get) if isinstance(d, dict) and d else "-"

    # ================= KPI =================
    total = summary.get("total_attacks", 0)
    top_country = get_top_attacker_country(summary.get("top_countries", {}))
    top_protocol = get_top(summary.get("protocol_dist", {}))
    peak_hour = get_top(summary.get("attacks_by_hour", {}))

    # ================= EXEC SUMMARY =================
    content.append(Paragraph("<b>Executive Summary</b>", styles["Heading2"]))
    content.append(Paragraph(
        f"A total of {total} suspicious events were detected. "
        f"Primary external origin: {top_country}. "
        f"Most abused protocol: {top_protocol}.",
        styles["Normal"]
    ))
    content.append(Spacer(1, 12))

    # ================= KPI TABLE =================
    kpi_data = [
        ["Metric", "Value"],
        ["Total Attacks", total],
        ["Top Attacker Country", top_country],
        ["Top Protocol", top_protocol],
        ["Peak Hour", peak_hour],
        ["Unique IPs", len(summary.get("top_ips", {}))],
        ["Targeted Units", len(summary.get("top_unit", {}))]
    ]

    table = Table(kpi_data, colWidths=[200, 150])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
    ]))

    content.append(table)
    content.append(Spacer(1, 15))

    # ================= CHARTS =================
    try:

        img_elements = []

        # Hourly
        hours = summary.get("attacks_by_hour", {})
        if hours:
            x = sorted(map(int, hours.keys()))
            y = [hours.get(str(i), 0) for i in x]

            plt.figure(figsize=(4,2))
            plt.bar(x, y)
            plt.title("Attacks by Hour")

            img1 = pdf_dir / "hourly.png"
            plt.savefig(img1, bbox_inches="tight")
            plt.close()

            img_elements.append(Image(str(img1), width=240, height=130))

        # Protocol (clean pie)
        proto = summary.get("protocol_dist", {})
        if proto and sum(proto.values()) > 0:

            plt.figure(figsize=(4,2))
            plt.pie(
                proto.values(),
                labels=proto.keys(),
                autopct="%1.1f%%",
                textprops={'fontsize': 6}
            )
            plt.title("Protocol")

            img2 = pdf_dir / "protocol.png"
            plt.savefig(img2, bbox_inches="tight")
            plt.close()

            img_elements.append(Image(str(img2), width=240, height=130))

        # Countries (exclude IN)
        countries = {
            k: v for k, v in summary.get("top_countries", {}).items()
            if k.upper() != "IN"
        }

        if countries:
            plt.figure(figsize=(4,2))
            plt.bar(list(countries.keys())[:10], list(countries.values())[:10])
            plt.title("Top Attacker Countries")

            img3 = pdf_dir / "countries.png"
            plt.savefig(img3, bbox_inches="tight")
            plt.close()

            img_elements.append(Image(str(img3), width=240, height=130))

        # IPs
        ips = summary.get("top_ips", {})
        if ips:
            plt.figure(figsize=(4,2))
            plt.barh(list(ips.keys())[:10], list(ips.values())[:10])
            plt.title("Top Attacker IPs")

            img4 = pdf_dir / "ips.png"
            plt.savefig(img4, bbox_inches="tight")
            plt.close()

            img_elements.append(Image(str(img4), width=240, height=130))

        # GRID (2x2)
        if len(img_elements) % 2 != 0:
            img_elements.append("")

        rows = [img_elements[i:i+2] for i in range(0, len(img_elements), 2)]

        grid = Table(rows, colWidths=[260, 260])
        grid.setStyle(TableStyle([
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))

        content.append(grid)
        content.append(Spacer(1, 15))

    except Exception as e:
        log(f"Chart error: {e}")

    # ================= TABLES =================
    def dict_to_table(title, data):
        if not data:
            return

        content.append(Paragraph(f"<b>{title}</b>", styles["Heading3"]))

        t_data = [["Value", "Count"]]
        for k, v in list(data.items())[:10]:
            t_data.append([str(k), str(v)])

        t = Table(t_data)
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.grey),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.5,colors.black),
        ]))

        content.append(t)
        content.append(Spacer(1, 10))

    dict_to_table("Top Source IPs", summary.get("top_ips"))
    dict_to_table("Top Ports", summary.get("top_dst_ports"))
    dict_to_table("Top Units", summary.get("top_unit"))

    # ================= INSIGHTS =================
    content.append(Paragraph("<b>Threat Analysis</b>", styles["Heading2"]))
    content.append(Paragraph(
        f"Traffic indicates automated intrusion patterns. "
        f"Peak activity observed at hour {peak_hour}. "
        f"External attack sources dominate over internal traffic.",
        styles["Normal"]
    ))

    # ================= RECOMMENDATIONS =================
    content.append(Paragraph("<b>Recommendations</b>", styles["Heading2"]))

    for r in [
        "Block malicious IPs",
        "Enable geo-blocking",
        "Harden exposed services",
        "Deploy IDS/IPS signatures",
        "Monitor internal lateral movement"
    ]:
        content.append(Paragraph(f"• {r}", styles["Normal"]))

    # ================= BUILD =================
    doc.build(content)

    log(f"PDF saved at: {pdf_path}")
    log("PDF REPORT GENERATED (FINAL)")


# ================= NAVY COMBINE =================
# ================= NAVY COMBINE =================
def combine_navy_ioc_auto(base_folder, max_rows=1_000_000):

    import re

    base_path = Path(base_folder)

    # ✅ ROOT IOC PATH (NO ASSUMPTION OF NAVY FOLDER)
    ioc_path = base_path / "output" / "ioc_hits_mail"

    if not ioc_path.exists():
        log("❌ IOC folder not found")
        return

    # ✅ OUTPUT PATH
    combined_path = base_path / "output" / "ioc_hits_mail" / "Combined" / "navy"
    combined_path.mkdir(parents=True, exist_ok=True)

    # ✅ GET NAVY FILES FROM ANYWHERE
    files = [
        f for f in ioc_path.rglob("*_ioc_hits_mail.csv")
        if "navy" in f.name.lower()
    ]

    if not files:
        log("❌ No Navy IOC files found for combine")
        return

    # ✅ SORT FILES
    def extract_num(f):
        m = re.search(r"result_file_(\d+)", f.name)
        return int(m.group(1)) if m else 0

    files = sorted(files, key=extract_num)

    all_df = []

    for file in files:
        log(f"[NAVY COMBINE] Reading {file.name}")

        try:
            df = pd.read_csv(file)
            df.columns = df.columns.str.strip()

            if "Time" not in df.columns:
                log(f"[NAVY COMBINE] Skipping (no Time): {file.name}")
                continue

            df["Time"] = pd.to_datetime(df["Time"], errors="coerce", dayfirst=True)
            df = df.dropna(subset=["Time"])

            all_df.append(df)

        except Exception as e:
            log(f"[NAVY COMBINE] Error: {file.name} → {e}")

    if not all_df:
        log("❌ No valid Navy data to combine")
        return

    final_df = pd.concat(all_df, ignore_index=True)
    final_df = final_df.sort_values("Time")

    total_rows = len(final_df)
    log(f"[NAVY COMBINE] Total rows: {total_rows}")

    file_index = 1

    for start in range(0, total_rows, max_rows):

        chunk = final_df.iloc[start:start + max_rows]

        output_file = combined_path / f"combined_navy_{file_index}.csv"
        chunk.to_csv(output_file, index=False)

        log(f"[NAVY COMBINE] Saved: {output_file} | Rows: {len(chunk)}")

        file_index += 1


# ================= AIRFORCE COMBINE =================
def combine_airforce_ioc_auto(base_folder, max_rows=1_000_000):

    import re

    base_path = Path(base_folder)

    # ✅ ROOT IOC PATH (NO ASSUMPTION OF AIRFORCE FOLDER)
    ioc_path = base_path / "output" / "ioc_hits_mail"

    if not ioc_path.exists():
        log("❌ IOC folder not found")
        return

    # ✅ OUTPUT PATH
    combined_path = base_path / "output" / "ioc_hits_mail" / "Combined" / "airforce"
    combined_path.mkdir(parents=True, exist_ok=True)

    # ✅ GET AIRFORCE FILES FROM ANYWHERE
    files = [
        f for f in ioc_path.rglob("*_ioc_hits_mail.csv")
        if "airforce" in f.name.lower() or "air force" in f.name.lower()
    ]

    if not files:
        log("❌ No Airforce IOC files found for combine")
        return

    # ✅ SORT FILES
    def extract_num(f):
        m = re.search(r"result_file_(\d+)", f.name)
        return int(m.group(1)) if m else 0

    files = sorted(files, key=extract_num)

    all_df = []

    for file in files:
        log(f"[AIRFORCE COMBINE] Reading {file.name}")

        try:
            df = pd.read_csv(file)
            df.columns = df.columns.str.strip()

            if "Time" not in df.columns:
                log(f"[AIRFORCE COMBINE] Skipping (no Time): {file.name}")
                continue

            df["Time"] = pd.to_datetime(df["Time"], errors="coerce", dayfirst=True)
            df = df.dropna(subset=["Time"])

            all_df.append(df)

        except Exception as e:
            log(f"[AIRFORCE COMBINE] Error: {file.name} → {e}")

    if not all_df:
        log("❌ No valid Airforce data to combine")
        return

    final_df = pd.concat(all_df, ignore_index=True)
    final_df = final_df.sort_values("Time")

    total_rows = len(final_df)
    log(f"[AIRFORCE COMBINE] Total rows: {total_rows}")

    file_index = 1

    for start in range(0, total_rows, max_rows):

        chunk = final_df.iloc[start:start + max_rows]

        output_file = combined_path / f"combined_airforce_{file_index}.csv"
        chunk.to_csv(output_file, index=False)

        log(f"[AIRFORCE COMBINE] Saved: {output_file} | Rows: {len(chunk)}")

        file_index += 1
# ================= MAIN =================
def process(folder):

    progress["done"] = False
    progress["folder"] = folder
    progress["pdf_path"] = None  # reset for download link

    path = Path(folder)

    output_base = path / "output"

    cleaned = output_base / "csv_cleaned"
    ioc_dir = output_base / "ioc_hits_mail"
    reports = output_base / "reports"

    json_dir = reports/"json"
    dcya_dir = reports/"dcya"
    pdf_dir = reports/"pdf"

    cleaned.mkdir(parents=True, exist_ok=True)
    ioc_dir.mkdir(parents=True, exist_ok=True)
    reports.mkdir(parents=True, exist_ok=True)

    log("START")

    # 🔥 CLEAN FIXED
    log("Cleaning started")
    for f in path.rglob("*.csv"):

        if "output" in str(f).lower():
            continue

        if not is_valid_file(f):
            continue

        log(f"Cleaning: {f.name}")
        rel = f.relative_to(path)
        clean_csv(f, cleaned/rel)

    # 🔥 MYSQL
    ioc_set, unit_map = load_reference_data()

    # 🔥 MATCH
    log("IOC matching started")
    for f in cleaned.rglob("*.csv"):
        hits = process_file(f, ioc_set, unit_map, ioc_dir)
        log(f"{f.name} → {hits}")

    time.sleep(1)

    # 🔥 SUMMARY
    log("Generating summary")
    build_summary(ioc_dir, json_dir)

    # 🔥 NAVY COMBINE
    log("Running Navy Combine")
    combine_navy_ioc_auto(folder)
    log("Navy Combine Done")

    # AIRFORCE COMBINE
    log("Running Airforce Combine")
    combine_airforce_ioc_auto(folder)
    log("Airforce Combine Done")
    
    # 🔥 DCYA
    try:
        log("Generating DCYA report")
        build_dcya_report(ioc_dir, dcya_dir)
    except Exception as e:
        log(f"DCYA ERROR: {e}")

    # 🔥 PDF (executive version, falls back to basic automatically)
    try:
        log("Generating PDF report")
        build_pdf_report(ioc_dir, pdf_dir, json_dir)
        progress["pdf_path"] = str(pdf_dir / "SOC_REPORT_EXEC.pdf")
    except Exception as e:
        log(f"PDF ERROR: {e}")

    progress["status"] = "completed"
    log("DONE")
